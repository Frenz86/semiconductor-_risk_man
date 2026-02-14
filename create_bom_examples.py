"""
Script per creare 2 file BOM di esempio realistici e aggiornare il database.
- 02_BOM_Automotive_ADAS_ECU_15.xlsx (15 componenti - ECU ADAS automotive)
- 03_BOM_Industrial_IoT_Gateway_12.xlsx (12 componenti - Gateway IoT industriale)
"""

import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime

# =========================================================================
# COLUMN HEADERS (identici al template esistente)
# =========================================================================

HEADERS = [
    'Supplier Name',
    'Supplier Part Number',
    'Stand-Alone Functional Device (Y/N)',
    'In case answer on Column C is Y, Which other device in the BOM is necessary to run the PN on Column B? (e.g. PMIC for MPU, Memory for MPU)',
    'Unit Price ($)',
    'How Many Device of this specific PN are in the BOM?',
    'Category of product (MCU, MPU, Sensor, Analogic, Power, Passive Component, Transceiver Wireless)',
    'Commodity (Y/N)*',
    'Supplier Lead Time (weeks)',
    'Proprietary (Y/N)**',
    'Size of SW Code/Memory used (Only Semiconductor)',
    'Embedded Memory (Y/N)',
    'External Memory (Yes, No)',
    'Operating System Used (Only Semiconductor: None, Proprietary, Windows, Linux, Zephir...)',
    'Wireless Protocol Used (Only Semiconductor: WiFi, BLE, Matter, etc\u2026)',
    'Specify Certification/Qualification',
    'Weeks to qualify',
    'Technology Node (Semiconductor Component)',
    'Manufacturing Plant 1 (Front-End, Back-End, Test and Assembly, Passive Production)',
    'Country of Manufacturing Plant 1',
    'Manufacturing Plant 2 (Front-End, Back-End, Test and Assembly, Passive Production)',
    'Country of Manufacturing Plant 2',
    'Manufacturing Plant 3 (Front-End, Back-End, Test and Assembly, Passive Production)',
    'Country of Manufacturing Plant 3',
    'Manufacturing Plant 4 (Front-End, Back-End, Test and Assembly, Passive Production)',
    'Country of Manufacturing Plant 4',
    'BOM/PCB Produced by EMS?',
    'EMS (if yes) name',
    'EMS (if yes) location',
    'Inventory/Unit (availability) to EMS',
    'Does EMS by from Distributor or directly from Supplier?',
    'Name of Distributor',
    'Location of Distributor',
    'Dedicated Inventory/Unit (availability) to the distributor',
    'Dedicated Buffer Stock Units to the supplier',
    'If Dedicated Buffer Stock Units to the supplier is yes specify the number of Units',
    'EOL_Status\n(Active/NRND/Last_Buy/EOL/Obsolete)',
    'Number_of_Alternative_Sources\n(0, 1, 2, 3+)',
    'Supplier_Financial_Health\n(A/B/C/D)',
    'MTBF_Hours',
    'Automotive_Grade\n(None/AEC-Q100/Q101/Q200)',
    'Last_Price_Increase_Pct\n(%)',
    'Allocation_Status\n(Normal/Constrained/Allocated)',
    'Package_Type\n(QFP/BGA/WLCSP/QFN/SOP/DIP)',
]

ABSTRACT_ROWS = [
    ['Supply Chain Risk Assessment - Input Template v3.1'],
    [],
    ['The electronic supply chain is one of the most complex in the world. '
     'This tool assesses BOM-level risk and provides resilience suggestions. '
     'The INPUTS sheet contains all the data fields described below. '
     'Fields marked as NEW v3.1 are optional - if left empty, safe defaults are used.'],
    [],
    ['#', 'Campo (Column Name)', 'Descrizione', 'Valori Ammessi', 'Impatto Score'],
    ['IDENTIFICAZIONE COMPONENTE'],
    ['1', 'Supplier Name', 'Nome del fornitore/produttore del componente',
     'Testo libero (es. NXP, STMicroelectronics)', '-'],
    ['2', 'Supplier Part Number', 'Codice univoco del componente (Part Number del fornitore)',
     'Testo libero (es. STM32MP157CAC3)', '-'],
    ['3', 'Category of product', 'Categoria merceologica del componente',
     'MCU, MPU, Sensor, Analogic, Power, Passive Component, Transceiver Wireless, Memory, Connector', '-'],
    ['4', 'Unit Price (USD)', 'Prezzo unitario del componente in dollari',
     'Numero (es. 12.50)', 'Peso finanziario per media BOM'],
    ['5', 'How Many Device in the BOM?', 'Quantita del componente utilizzata in ogni singola PCB/BOM',
     'Numero intero (es. 1, 2, 45)', 'Peso finanziario per media BOM'],
    ['DIPENDENZE FUNZIONALI'],
    ['6', 'Stand-Alone Functional Device (Y/N)', 'Il componente funziona autonomamente o dipende da altri nella BOM?',
     'Y = standalone, N = dipende da altri', '+10 punti se N'],
    ['7', 'Which other device is necessary?', 'Se non standalone, da quali componenti dipende (es. MPU dipende da PMIC e DDR)',
     'Testo libero (es. PMIC STPMIC1, DDR Memory)', 'Chain risk propagation'],
    ['8', 'Commodity (Y/N)*', 'Componente facilmente reperibile con alternative dirette sul mercato?',
     'Y = commodity, N = specifico', '+5 punti se N'],
    ['9', 'Proprietary (Y/N)**', 'Componente proprietario senza alternative pin-to-pin?',
     'Y = proprietario, N = standard', '+10 punti se Y'],
    ['SUPPLY CHAIN E LEAD TIME'],
    ['10', 'Supplier Lead Time (weeks)', 'Tempo di consegna standard del fornitore in settimane',
     'Numero (es. 18)', 'Fino a +15 punti (>26w = critico)'],
    ['11', 'Buffer Stock Units', 'Numero di unita in buffer stock dedicato presso il fornitore',
     'Numero (es. 50000)', 'Fino a +15 punti o -5 bonus'],
    ['SOFTWARE / FIRMWARE'],
    ['12', 'Size of SW Code/Memory used (KB)', 'Dimensione del codice SW/firmware in KB (solo semiconduttori)',
     'Numero in KB (es. 2048)', 'Calcolo switching cost'],
    ['13', 'Embedded Memory (Y/N)', 'Il componente ha memoria integrata?',
     'Y / N', 'Informativo'],
    ['14', 'External Memory (Yes, No)', 'Il componente usa memoria esterna?',
     'Yes / No', 'Informativo'],
    ['15', 'Operating System Used', 'Sistema operativo utilizzato (solo semiconduttori attivi)',
     'None, Baremetal, Proprietary, FreeRTOS, Linux, Windows, Zephyr', 'Calcolo switching cost (moltiplicatore)'],
    ['16', 'Wireless Protocol Used', 'Protocollo wireless utilizzato (solo semiconduttori)',
     'WiFi, BLE, Matter, Zigbee, Thread, LoRa, NB-IoT', 'Calcolo switching cost'],
    ['CERTIFICAZIONI E QUALIFICA'],
    ['17', 'Specify Certification/Qualification', 'Certificazioni richieste per il componente',
     'Testo (es. AEC-Q100, VDE-Class C, MIL-STD)', '+5 punti se riqualifica > 12 sett.'],
    ['18', 'Weeks to qualify', 'Settimane necessarie per qualificare un componente alternativo',
     'Numero (es. 24)', '+5 punti se > 12 settimane'],
    ['PRODUZIONE E GEOGRAFIA'],
    ['19', 'Technology Node', 'Nodo tecnologico del semiconduttore (dimensione minima del transistor)',
     'Testo (es. 28nm, 7nm, 130nm, MEMS)', '+3/+5 punti per nodi avanzati'],
    ['20', 'Manufacturing Plant 1 (Name)', 'Nome dello stabilimento produttivo 1 (Front-End, Back-End, Test)',
     'Testo (es. TSMC, Samsung)', '-'],
    ['21', 'Country of Manufacturing Plant 1', 'Paese dello stabilimento produttivo 1',
     'Nome paese (es. Taiwan, France, USA)', 'Fino a +25 punti (concentraz. geo)'],
    ['22', 'Manufacturing Plant 2-4 + Country', 'Stabilimenti produttivi aggiuntivi (fino a 4 plant)',
     'Come sopra', 'Piu plant = meno rischio'],
    ['EMS E DISTRIBUZIONE'],
    ['23', 'BOM/PCB Produced by EMS?', 'La PCB e prodotta da un assemblatore esterno (EMS)?',
     'Yes / No', 'Informativo'],
    ['24', 'EMS name / location', 'Nome e sede dell EMS utilizzato',
     'Testo (es. Jabil, Poland)', 'Informativo'],
    ['25', 'Inventory/Unit to EMS', 'Unita disponibili presso EMS',
     'Numero', 'Informativo'],
    ['26', 'Distributor name / location', 'Distributore utilizzato e sua sede',
     'Testo (es. Arrow, Germany)', 'Informativo'],
    ['27', 'Dedicated Inventory to distributor', 'Unita dedicate disponibili presso il distributore',
     'Numero', 'Informativo'],
    ['EXTENDED RISK ASSESSMENT v3.1 (NUOVI)'],
    ['28', 'EOL_Status',
     'Stato del ciclo di vita del componente. EOL/Obsolete = fine produzione annunciata. NRND = non raccomandato per nuovi design. Last_Buy = ultima opportunita di acquisto prima della discontinuazione.',
     'Active, NRND, Last_Buy, EOL, Obsolete', 'Fino a +15 (EOL = critico)'],
    ['29', 'Number_of_Alternative_Sources',
     'Quante fonti alternative (second source pin-to-pin o funzionale) esistono sul mercato, indipendentemente dal numero di stabilimenti del fornitore attuale.',
     '0, 1, 2, 3, 4, 5+', '+10 (0 fonti), +5 (1), -3 bonus (3+)'],
    ['30', 'Supplier_Financial_Health',
     'Rating salute finanziaria del fornitore. A = solido, fatturato stabile. B = leggere criticita. C = rischio medio, possibile ristrutturazione. D = rischio insolvenza, acquisizione ostile o crisi.',
     'A, B, C, D', 'Fino a +8 (D = critico)'],
    ['31', 'MTBF_Hours',
     'Mean Time Between Failures - affidabilita del componente in ore. Valori bassi indicano rischio di guasti frequenti e necessita di scorte maggiori di ricambio.',
     'Numero ore (es. 500000)', 'Informativo (segnala se < 50.000h)'],
    ['32', 'Automotive_Grade',
     'Grado di qualifica automotive del componente. I componenti AEC-Q hanno supply chain piu rigide, lead time piu lunghi, ma garanzia di affidabilita superiore.',
     'None, AEC-Q100, AEC-Q101, AEC-Q200', 'Informativo + switching cost'],
    ['33', 'Last_Price_Increase_Pct',
     'Percentuale dell ultimo aumento di prezzo applicato dal fornitore. Aumenti significativi (>20%) sono un segnale di tensione nella supply chain o di posizione dominante del fornitore.',
     'Numero % (es. 15, 50)', '+5 se >50%, +3 se >20%'],
    ['34', 'Allocation_Status',
     'Stato attuale di allocazione. Normal = disponibilita regolare. Constrained = disponibilita ridotta con lead time allungati. Allocated = forniture contingentate dal fornitore.',
     'Normal, Constrained, Allocated', '+10 (Allocated), +5 (Constrained)'],
    ['35', 'Package_Type',
     'Tipo di package del componente. Package avanzati (WLCSP, FCBGA, FOWLP) sono prodotti da poche fonderie specializzate, aumentando il rischio di concentrazione produttiva.',
     'QFP, BGA, WLCSP, FCBGA, QFN, SOP, DIP, CSP, SOT-23, Module', '+3 per package avanzati'],
    ['NOTE:'],
    ['* Commodity: componente con alternative dirette disponibili sul mercato (drop-in replacement)'],
    ['** Proprietary: componente con design proprietario del fornitore, senza alternative pin-to-pin compatibili'],
    ['I campi v3.1 (evidenziati in giallo) sono opzionali. Se lasciati vuoti vengono usati i default: Active, Normal, A.'],
    ['Lo score di rischio totale e cappato a 100 punti. Score >= 55 = ALTO (rosso), 30-54 = MEDIO (giallo), < 30 = BASSO (verde).'],
]

OUTPUT_HEADERS = ['Color Code', 'Risk Description', 'Suggestion', 'Man Hour Impact']


def create_bom_file(filename, run_rate, components, board_title):
    """Crea un file BOM Excel con la struttura standard."""
    wb = openpyxl.Workbook()

    # Elimina il foglio di default
    wb.remove(wb.active)

    # --- ABSTRACT SHEET ---
    ws_abstract = wb.create_sheet('Abstract', 0)

    # Stili
    title_font = Font(bold=True, size=14)
    title_align = Alignment(horizontal='center', vertical='center')
    desc_align = Alignment(wrap_text=True, vertical='top')
    header_font = Font(bold=True)
    category_font = Font(bold=True)
    category_fill = PatternFill(start_color='B4C6E7', end_color='B4C6E7', fill_type='solid')
    header_fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
    row_fill_even = PatternFill(start_color='F2F2F2', end_color='F2F2F2', fill_type='solid')

    # Titolo principale (riga 1)
    ws_abstract.merge_cells('A1:E1')
    c1 = ws_abstract['A1']
    c1.value = 'Supply Chain Risk Assessment - Input Template v3.1'
    c1.font = title_font
    c1.alignment = title_align
    ws_abstract.row_dimensions[1].height = 25

    # Descrizione (riga 3)
    ws_abstract.merge_cells('A3:E3')
    c3 = ws_abstract['A3']
    c3.value = ('The electronic supply chain is one of the most complex in the world. '
                'This tool assesses BOM-level risk and provides resilience suggestions. '
                'The INPUTS sheet contains all the data fields described below. '
                'Fields marked as NEW v3.1 are optional - if left empty, safe defaults are used.')
    c3.alignment = desc_align
    ws_abstract.row_dimensions[3].height = 40

    # Header tabella (riga 5)
    headers = ['#', 'Campo (Column Name)', 'Descrizione', 'Valori Ammessi', 'Impatto Score']
    for col_idx, val in enumerate(headers, start=1):
        cell = ws_abstract.cell(row=5, column=col_idx, value=val)
        cell.font = header_font
        cell.fill = header_fill

    # Categorie e campi (partendo dalla riga 6)
    # Salta i primi 4 elementi di ABSTRACT_ROWS (title vuoto, desc, vuoto, header)
    # che abbiamo già scritto manualmente
    row_idx = 6
    content_rows = ABSTRACT_ROWS[5:]  # Salta le prime 5 righe

    for row_data in content_rows:
        if len(row_data) == 1 and row_data[0]:
            # Categoria - merge celle
            ws_abstract.merge_cells(f'A{row_idx}:E{row_idx}')
            cell = ws_abstract.cell(row=row_idx, column=1, value=row_data[0])
            cell.font = category_font
            cell.fill = category_fill
            row_idx += 1
        elif len(row_data) >= 2:
            # Campo normale
            for col_idx, val in enumerate(row_data[:5], start=1):
                cell = ws_abstract.cell(row=row_idx, column=col_idx, value=val)
                # Alterna colori righe
                if row_idx % 2 == 1:
                    cell.fill = row_fill_even
            row_idx += 1

    # Larghezza colonne Abstract
    ws_abstract.column_dimensions['A'].width = 5
    ws_abstract.column_dimensions['B'].width = 18
    ws_abstract.column_dimensions['C'].width = 50
    ws_abstract.column_dimensions['D'].width = 35
    ws_abstract.column_dimensions['E'].width = 20

    # --- INPUTS SHEET ---
    ws_inputs = wb.create_sheet('INPUTS')

    # Row 1: Board title
    ws_inputs.append([board_title])
    ws_inputs['A1'].font = Font(bold=True, size=14, color='1F4E79')

    # Row 2: empty
    ws_inputs.append([])

    # Row 3: Run Rate + section headers
    row3 = ['Run Rate (Number of PCB per week)', run_rate]
    while len(row3) < 10:
        row3.append(None)
    row3.append('SW/FIRMWARE')
    while len(row3) < 15:
        row3.append(None)
    row3.append('CERTIFIED/QUALIFICATION BOM (Safety/SIL-ASIL, CRA, VDE, FDA, USA...)')
    while len(row3) < 36:
        row3.append(None)
    row3.append('EXTENDED RISK ASSESSMENT v3.1')
    ws_inputs.append(row3)

    # Row 4: empty
    ws_inputs.append([])

    # Row 5: Headers
    ws_inputs.append(HEADERS)
    header_row = ws_inputs[5]
    header_fill = PatternFill(start_color='1F4E79', end_color='1F4E79', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF', size=9)
    for cell in header_row:
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(wrap_text=True, vertical='center')

    # Data rows
    for comp in components:
        ws_inputs.append(comp)

    # Set column widths
    for col_idx in range(1, len(HEADERS) + 1):
        ws_inputs.column_dimensions[get_column_letter(col_idx)].width = 18
    ws_inputs.column_dimensions['A'].width = 22
    ws_inputs.column_dimensions['B'].width = 28
    ws_inputs.column_dimensions['D'].width = 35

    # --- OUTPUT SHEET ---
    ws_output = wb.create_sheet('OUTPUT')
    ws_output.append(OUTPUT_HEADERS)
    ws_output.append(['Overall BOM Risk (Red-High; Yellow-Mid; Green-Low)'])
    for i in range(len(components)):
        ws_output.append([f'Component {i+1}'])

    wb.save(filename)
    print(f"  Created: {filename} ({len(components)} components, run_rate={run_rate})")


# =========================================================================
# BOARD 2: AUTOMOTIVE ADAS ECU - 15 Components
# =========================================================================
# ECU per sistema ADAS (Advanced Driver Assistance System) con camera
# Rischi chiave: sole-source vision processor, NRND connector, advanced nodes
# EMS: Bosch Germany (tier-1 automotive)

adas_components = [
    # 1. NXP S32K344 - Automotive Safety MCU (ASIL-D)
    ['NXP', 'S32K344', 'Y', None,
     8.50, 1, 'MCU', 'N', 22, 'N',
     1024, 'Y', 'No', 'Baremetal', None,
     'AEC-Q100, ISO 26262 ASIL-D', 28, '40nm',
     'TSMC', 'Taiwan', 'NXP (ATMC)', 'Malaysia', 'NXP', 'Thailand', None, None,
     'Yes', 'Bosch', 'Germany', 15000,
     'Distributor', 'Arrow', 'Germany', 30000,
     'Yes', 75000,
     'Active', 2, 'A', 500000, 'AEC-Q100', 8, 'Normal', 'BGA'],

    # 2. NXP S32G274A - Vehicle Network Processor (sole source!)
    ['NXP', 'S32G274A', 'N', 'PMIC PF5024, LPDDR4 MT53E256M32D1DS-046',
     28.00, 1, 'MPU', 'N', 26, 'Y',
     4096, 'N', 'Yes', 'Linux', None,
     'AEC-Q100, ISO 26262 ASIL-B', 36, '16nm',
     'TSMC', 'Taiwan', 'JCET', 'China', None, None, None, None,
     'Yes', 'Bosch', 'Germany', 3000,
     'Supplier', None, None, None,
     'No', 0,
     'Active', 0, 'A', 400000, 'AEC-Q100', 15, 'Constrained', 'FCBGA'],

    # 3. NXP PF5024 - Safety PMIC (companion chip - sole source!)
    ['NXP', 'PF5024', 'N', 'S32G274A MPU',
     5.80, 1, 'Power', 'N', 20, 'Y',
     0, 'N', 'No', None, None,
     'AEC-Q100', 24, '130nm',
     'NXP', 'USA', 'NXP (ATMC)', 'Malaysia', None, None, None, None,
     'Yes', 'Bosch', 'Germany', 5000,
     'Supplier', None, None, None,
     'Yes', 15000,
     'Active', 0, 'A', 600000, 'AEC-Q100', 5, 'Normal', 'QFN'],

    # 4. Mobileye EyeQ5H - Vision Processor (7nm, sole source, allocated!)
    ['Mobileye (Intel)', 'EyeQ5H', 'N', 'LPDDR4 MT53E256M32D1DS-046, NOR Flash MX25U25645G',
     55.00, 1, 'MPU', 'N', 32, 'Y',
     8192, 'N', 'Yes', 'Linux', None,
     'AEC-Q100, ISO 26262 ASIL-B', 48, '7nm',
     'TSMC', 'Taiwan', 'ASE', 'Taiwan', None, None, None, None,
     'Yes', 'Bosch', 'Germany', 1000,
     'Supplier', None, None, None,
     'No', 0,
     'Active', 0, 'A', 350000, 'AEC-Q100', 0, 'Allocated', 'FCBGA'],

    # 5. Micron MT53E256M32D1DS-046 - LPDDR4 (shared by 2 processors)
    ['Micron', 'MT53E256M32D1DS-046', 'N', 'S32G274A MPU, EyeQ5H Vision Processor',
     6.50, 4, 'Memory', 'N', 18, 'N',
     0, 'N', 'Yes', None, None,
     'AEC-Q100', 16, '12nm',
     'Micron', 'Japan', 'Micron', 'China', None, None, None, None,
     'Yes', 'Bosch', 'Germany', 8000,
     'Distributor', 'WPG', 'Taiwan', 15000,
     'No', 0,
     'Active', 2, 'A', 200000, 'AEC-Q100', 30, 'Constrained', 'BGA'],

    # 6. Macronix MX25U25645G - NOR Flash 256Mbit
    ['Macronix', 'MX25U25645G', 'Y', None,
     2.80, 2, 'Memory', 'N', 14, 'N',
     0, 'N', 'No', None, None,
     'AEC-Q100', 12, '45nm',
     'Macronix', 'Taiwan', 'Macronix', 'Taiwan', None, None, None, None,
     'Yes', 'Bosch', 'Germany', 12000,
     'Distributor', 'Mouser', 'USA', 20000,
     'Yes', 40000,
     'Active', 3, 'B', 300000, 'AEC-Q100', 10, 'Normal', 'SOP'],

    # 7. Marvell 88Q2112 - Automotive Ethernet PHY 100BASE-T1
    ['Marvell', '88Q2112-A2-NNP2I000', 'Y', None,
     4.20, 2, 'Transceiver', 'N', 24, 'N',
     64, 'Y', 'No', 'Baremetal', None,
     'AEC-Q100', 20, '28nm',
     'TSMC', 'Taiwan', 'ASE', 'Malaysia', None, None, None, None,
     'Yes', 'Bosch', 'Germany', 6000,
     'Distributor', 'Arrow', 'Germany', 10000,
     'Yes', 25000,
     'Active', 1, 'A', 450000, 'AEC-Q100', 12, 'Constrained', 'QFN'],

    # 8. TI TPS65263-1QRGZRQ1 - Multi-output DC/DC Converter
    ['Texas Instruments', 'TPS65263-1QRGZRQ1', 'Y', None,
     3.50, 1, 'Power', 'Y', 12, 'N',
     0, 'N', 'No', None, None,
     'AEC-Q100', 10, '65nm',
     'TI', 'USA', 'TI', 'Germany', 'TSMC', 'Taiwan', None, None,
     'Yes', 'Bosch', 'Germany', 20000,
     'Distributor', 'Arrow', 'Germany', 50000,
     'Yes', 80000,
     'Active', 3, 'A', 700000, 'AEC-Q100', 3, 'Normal', 'QFN'],

    # 9. Bosch SMI230 - 6-axis IMU MEMS
    ['Bosch Sensortec', 'SMI230', 'Y', None,
     3.80, 1, 'Sensor', 'N', 16, 'N',
     0, 'N', 'No', None, None,
     'AEC-Q100', 14, 'MEMS',
     'Bosch', 'Germany', 'Bosch', 'China', None, None, None, None,
     'Yes', 'Bosch', 'Germany', 10000,
     'Supplier', None, None, None,
     'Yes', 30000,
     'Active', 2, 'A', 350000, 'AEC-Q100', 0, 'Normal', 'LGA'],

    # 10. NXP TJA1463 - CAN FD System Basis Chip
    ['NXP', 'TJA1463', 'Y', None,
     2.90, 3, 'Transceiver', 'N', 18, 'N',
     0, 'N', 'No', None, None,
     'AEC-Q100', 16, '130nm',
     'NXP', 'Germany', 'NXP (ATMC)', 'Malaysia', None, None, None, None,
     'Yes', 'Bosch', 'Germany', 15000,
     'Distributor', 'Avnet', 'Germany', 40000,
     'Yes', 60000,
     'Active', 1, 'A', 600000, 'AEC-Q100', 5, 'Normal', 'SOP'],

    # 11. Murata GCM32ER71E106KA37 - MLCC 10uF Automotive
    ['Murata', 'GCM32ER71E106KA37', 'Y', None,
     0.08, 65, 'Passive Component', 'Y', 8, 'N',
     0, 'N', 'No', None, None,
     'AEC-Q200', 4, None,
     'Murata', 'Japan', 'Murata', 'Philippines', 'Murata', 'Thailand', None, None,
     'Yes', 'Bosch', 'Germany', 500000,
     'Distributor', 'DigiKey', 'USA', 2000000,
     'Yes', 5000000,
     'Active', 4, 'A', 1000000, 'AEC-Q200', 0, 'Normal', '1206'],

    # 12. Nexperia PESD2CAN-U - CAN Bus TVS Protection
    ['Nexperia', 'PESD2CAN-U', 'Y', None,
     0.12, 6, 'Passive Component', 'Y', 6, 'N',
     0, 'N', 'No', None, None,
     'AEC-Q101', 3, None,
     'Nexperia', 'Germany', 'Nexperia', 'Malaysia', 'Nexperia', 'Philippines', None, None,
     'Yes', 'Bosch', 'Germany', 100000,
     'Distributor', 'Mouser', 'USA', 300000,
     'Yes', 500000,
     'Active', 3, 'A', 800000, 'AEC-Q101', 0, 'Normal', 'SOT-23'],

    # 13. TDK VLS252015HBX-1R0M - Power Inductor 1uH
    ['TDK', 'VLS252015HBX-1R0M', 'Y', None,
     0.15, 8, 'Passive Component', 'Y', 8, 'N',
     0, 'N', 'No', None, None,
     'AEC-Q200', 3, None,
     'TDK', 'Japan', 'TDK', 'Thailand', None, None, None, None,
     'Yes', 'Bosch', 'Germany', 200000,
     'Distributor', 'DigiKey', 'USA', 500000,
     'Yes', 800000,
     'Active', 5, 'A', 900000, 'AEC-Q200', 0, 'Normal', 'SMD'],

    # 14. Molex 5025781070 - Board-to-Board Connector
    ['Molex', '5025781070', 'Y', None,
     0.45, 4, 'Connector', 'Y', 6, 'N',
     0, 'N', 'No', None, None,
     None, 2, None,
     'Molex', 'Japan', 'Molex', 'China', None, None, None, None,
     'Yes', 'Bosch', 'Germany', 50000,
     'Distributor', 'DigiKey', 'USA', 150000,
     'Yes', 200000,
     'Active', 2, 'A', 500000, None, 0, 'Normal', 'Connector'],

    # 15. TE 1-2141530-1-AUT - Automotive MQS Header (NRND!)
    ['TE Connectivity', '1-2141530-1-AUT', 'Y', None,
     0.85, 2, 'Connector', 'Y', 8, 'N',
     0, 'N', 'No', None, None,
     'AEC-Q200', 4, None,
     'TE', 'Mexico', 'TE', 'China', 'TE', 'Germany', 'TE', 'USA',
     'Yes', 'Bosch', 'Germany', 60000,
     'Distributor', 'Arrow', 'Germany', 100000,
     'Yes', 200000,
     'NRND', 1, 'A', 600000, 'AEC-Q200', 15, 'Normal', 'Connector'],
]


# =========================================================================
# BOARD 3: INDUSTRIAL IoT EDGE GATEWAY - 12 Components
# =========================================================================
# Gateway IoT per fabbrica intelligente (Industria 4.0)
# Rischi chiave: EOL sensor, Last_Buy flash, NRND WiFi, fornitore C-rating
# EMS: Foxconn China

iot_components = [
    # 1. Renesas RZ/G2L - MPU Linux (ARM Cortex-A55 + M33)
    ['Renesas', 'R9A07G044L23GBG', 'N', 'PMIC RAA215300, DDR3L AS4C256M16D3C-12',
     15.00, 1, 'MPU', 'N', 20, 'Y',
     4096, 'N', 'Yes', 'Linux', None,
     'IEC 62443', 24, '22nm',
     'TSMC', 'Taiwan', 'Amkor', 'Malaysia', None, None, None, None,
     'Yes', 'Foxconn', 'China', 5000,
     'Distributor', 'Avnet', 'Germany', 8000,
     'No', 0,
     'Active', 1, 'A', 350000, None, 10, 'Normal', 'BGA'],

    # 2. Renesas RAA215300 - PMIC companion (sole source!)
    ['Renesas', 'RAA215300', 'N', 'R9A07G044L23GBG MPU',
     3.50, 1, 'Power', 'N', 18, 'Y',
     0, 'N', 'No', None, None,
     'IEC 62443', 20, '130nm',
     'Renesas', 'Japan', 'Renesas', 'Malaysia', None, None, None, None,
     'Yes', 'Foxconn', 'China', 8000,
     'Distributor', 'Avnet', 'Germany', 12000,
     'Yes', 20000,
     'Active', 0, 'A', 500000, None, 5, 'Normal', 'QFN'],

    # 3. Qualcomm QCA6174A-5 - WiFi 5 + BT 4.2 (NRND!)
    ['Qualcomm', 'QCA6174A-5', 'Y', None,
     8.90, 1, 'Transceiver Wireless', 'N', 22, 'Y',
     2048, 'N', 'No', 'Linux', 'WiFi, BLE',
     'FCC, CE, TELEC', 18, '28nm',
     'TSMC', 'Taiwan', 'ASE', 'China', None, None, None, None,
     'Yes', 'Foxconn', 'China', 3000,
     'Distributor', 'Arrow', 'USA', 5000,
     'No', 0,
     'NRND', 1, 'A', 300000, None, 25, 'Constrained', 'BGA'],

    # 4. Quectel EC25-E - LTE Cat 4 Module
    ['Quectel', 'EC25-E', 'Y', None,
     18.50, 1, 'Transceiver Wireless', 'N', 14, 'N',
     512, 'N', 'No', 'Linux', 'LTE',
     'CE, FCC, PTCRB', 16, '28nm',
     'Qualcomm', 'China', 'Quectel', 'China', None, None, None, None,
     'Yes', 'Foxconn', 'China', 6000,
     'Distributor', 'Rutronik', 'Germany', 10000,
     'Yes', 15000,
     'Active', 2, 'B', 200000, None, 0, 'Normal', 'Module'],

    # 5. Micron MT29F4G08ABAFAWP - NAND Flash 4Gbit (LAST BUY!)
    ['Micron', 'MT29F4G08ABAFAWP', 'Y', None,
     3.20, 1, 'Memory', 'N', 16, 'N',
     0, 'N', 'No', None, None,
     None, 8, '14nm',
     'Micron', 'Japan', 'Micron', 'Singapore', None, None, None, None,
     'Yes', 'Foxconn', 'China', 4000,
     'Distributor', 'Mouser', 'USA', 6000,
     'No', 0,
     'Last_Buy', 2, 'A', 250000, None, 40, 'Allocated', 'TSOP'],

    # 6. Alliance Memory AS4C256M16D3C-12 - DDR3L (C-rated supplier!)
    ['Alliance Memory', 'AS4C256M16D3C-12', 'N', 'R9A07G044L23GBG MPU',
     2.80, 2, 'Memory', 'N', 14, 'N',
     0, 'N', 'Yes', None, None,
     None, 12, '25nm',
     'SMIC', 'China', 'JCET', 'China', None, None, None, None,
     'Yes', 'Foxconn', 'China', 10000,
     'Distributor', 'Mouser', 'USA', 20000,
     'Yes', 30000,
     'Active', 3, 'C', 200000, None, 18, 'Normal', 'BGA'],

    # 7. Microchip LAN9250 - Dual Ethernet 10/100 (long lead, price spike!)
    ['Microchip', 'LAN9250', 'Y', None,
     4.50, 1, 'Transceiver', 'N', 28, 'N',
     256, 'Y', 'No', 'Linux', None,
     'IEC 62443', 14, '90nm',
     'Microchip', 'USA', 'Amkor', 'Thailand', None, None, None, None,
     'Yes', 'Foxconn', 'China', 3000,
     'Distributor', 'DigiKey', 'USA', 5000,
     'Yes', 8000,
     'Active', 1, 'A', 400000, None, 55, 'Allocated', 'QFP'],

    # 8. Silicon Labs Si7021-A20 - Temp/Humidity Sensor (EOL!)
    ['Silicon Labs', 'Si7021-A20-GM1R', 'Y', None,
     2.10, 1, 'Sensor', 'Y', 10, 'N',
     0, 'N', 'No', None, None,
     None, 6, '180nm',
     'Silicon Labs', 'USA', 'ASE', 'Malaysia', None, None, None, None,
     'Yes', 'Foxconn', 'China', 15000,
     'Distributor', 'DigiKey', 'USA', 30000,
     'Yes', 50000,
     'EOL', 3, 'A', 300000, None, 0, 'Normal', 'DFN'],

    # 9. TDK C2012X7R1H104K - MLCC 100nF (low risk, commodity)
    ['TDK', 'C2012X7R1H104K', 'Y', None,
     0.01, 120, 'Passive Component', 'Y', 6, 'N',
     0, 'N', 'No', None, None,
     None, 2, None,
     'TDK', 'Japan', 'TDK', 'Philippines', 'TDK', 'China', None, None,
     'Yes', 'Foxconn', 'China', 1000000,
     'Distributor', 'DigiKey', 'USA', 5000000,
     'Yes', 10000000,
     'Active', 5, 'A', 1000000, None, 0, 'Normal', '0805'],

    # 10. Wurth 744043100 - Common Mode Choke 1mH
    ['Wurth Elektronik', '744043100', 'Y', None,
     0.65, 4, 'Passive Component', 'Y', 10, 'N',
     0, 'N', 'No', None, None,
     None, 4, None,
     'Wurth', 'Germany', 'Wurth', 'Germany', None, None, None, None,
     'Yes', 'Foxconn', 'China', 50000,
     'Distributor', 'Wurth Direct', 'Germany', 100000,
     'Yes', 150000,
     'Active', 3, 'A', 800000, None, 0, 'Normal', 'SMD'],

    # 11. Amphenol 10118194-0001LF - USB-C Connector
    ['Amphenol', '10118194-0001LF', 'Y', None,
     0.55, 2, 'Connector', 'Y', 8, 'N',
     0, 'N', 'No', None, None,
     None, 3, None,
     'Amphenol', 'China', 'Amphenol', 'China', None, None, None, None,
     'Yes', 'Foxconn', 'China', 30000,
     'Distributor', 'DigiKey', 'USA', 80000,
     'Yes', 100000,
     'Active', 4, 'A', 500000, None, 5, 'Normal', 'USB-C'],

    # 12. Vishay IHLP2525CZER1R0M11 - Power Inductor 1uH
    ['Vishay', 'IHLP2525CZER1R0M11', 'Y', None,
     0.35, 6, 'Passive Component', 'Y', 8, 'N',
     0, 'N', 'No', None, None,
     None, 3, None,
     'Vishay', 'USA', 'Vishay', 'Mexico', None, None, None, None,
     'Yes', 'Foxconn', 'China', 80000,
     'Distributor', 'Mouser', 'USA', 200000,
     'Yes', 300000,
     'Active', 4, 'A', 700000, None, 0, 'Normal', 'SMD'],
]


# =========================================================================
# MAIN: Create files
# =========================================================================

if __name__ == '__main__':
    print("Creating BOM example files...")
    print()

    create_bom_file(
        '02_BOM_Automotive_ADAS_ECU_15.xlsx',
        run_rate=3000,
        components=adas_components,
        board_title='AUTOMOTIVE ADAS ECU - Camera Control Module'
    )

    create_bom_file(
        '03_BOM_Industrial_IoT_Gateway_12.xlsx',
        run_rate=8000,
        components=iot_components,
        board_title='INDUSTRIAL IoT EDGE GATEWAY - Smart Factory Module'
    )

    print()
    print("Done! Both BOM files created successfully.")
